"""Professional PDF & Excel report generation for SmartDrive."""
from __future__ import annotations

import io
from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer, HRFlowable
)
from reportlab.lib.enums import TA_CENTER, TA_RIGHT


# ── Brand colours ─────────────────────────────────────────────────────────────
BRAND_BLUE  = colors.HexColor("#1a4d8c")
BRAND_GREEN = colors.HexColor("#16a34a")
BRAND_LIGHT = colors.HexColor("#e8f0fb")
GREY_HEADER = colors.HexColor("#f1f5f9")
GREY_TEXT   = colors.HexColor("#64748b")
WHITE       = colors.white


def _styles():
    base = getSampleStyleSheet()
    styles = {
        "title": ParagraphStyle(
            "SDTitle", parent=base["Normal"],
            fontSize=26, leading=32, textColor=BRAND_BLUE,
            spaceAfter=2, fontName="Helvetica-Bold",
        ),
        "subtitle": ParagraphStyle(
            "SDSubtitle", parent=base["Normal"],
            fontSize=11, textColor=GREY_TEXT, spaceAfter=4,
        ),
        "section": ParagraphStyle(
            "SDSection", parent=base["Normal"],
            fontSize=13, leading=18, textColor=BRAND_BLUE,
            spaceBefore=14, spaceAfter=6, fontName="Helvetica-Bold",
        ),
        "normal": ParagraphStyle(
            "SDNormal", parent=base["Normal"],
            fontSize=9, leading=13, textColor=colors.HexColor("#1e293b"),
        ),
        "footer": ParagraphStyle(
            "SDFooter", parent=base["Normal"],
            fontSize=8, textColor=GREY_TEXT, alignment=TA_CENTER,
        ),
    }
    return styles


def generate_booking_report(
    bookings: list,
    vehicles_map: dict,
    users_map: dict,
    start_label: str = "All Time",
    end_label: str = "Present",
) -> io.BytesIO:
    """
    Generate a print-ready / spiral-bound PDF booking report.
    Returns a BytesIO buffer ready for Flask's send_file.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=55, leftMargin=55,
        topMargin=60, bottomMargin=55,
        title="SmartDrive Bookings Report",
        author="SmartDrive Admin",
    )

    s = _styles()
    story = []

    # ── Header block ──────────────────────────────────────────────────────────
    story.append(Paragraph("SMARTDRIVE", s["title"]))
    story.append(Paragraph("Vehicle Hire &amp; Fleet Management Platform", s["subtitle"]))
    story.append(Paragraph(
        f"Bookings Report &nbsp;|&nbsp; {start_label} – {end_label} &nbsp;|&nbsp; "
        f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M EAT')}",
        s["subtitle"],
    ))
    story.append(HRFlowable(width="100%", thickness=2, color=BRAND_BLUE, spaceAfter=14))

    # ── Summary KPI table ─────────────────────────────────────────────────────
    total       = len(bookings)
    completed   = sum(1 for b in bookings if b.get("status") == "completed")
    approved    = sum(1 for b in bookings if b.get("status") == "approved")
    pending     = sum(1 for b in bookings if b.get("status") == "pending")
    rejected    = sum(1 for b in bookings if b.get("status") == "rejected")
    total_rev   = sum(b.get("total_amount", 0) for b in bookings
                      if b.get("payment_status") == "paid")
    paid_count  = sum(1 for b in bookings if b.get("payment_status") == "paid")

    kpi_data = [
        ["Total Bookings", "Revenue (KES)", "Completed", "Approved", "Pending", "Rejected"],
        [
            str(total),
            f"{total_rev:,.2f}",
            str(completed),
            str(approved),
            str(pending),
            str(rejected),
        ],
    ]
    kpi_col_w = [1.3 * inch] * 6
    kpi_table = Table(kpi_data, colWidths=kpi_col_w)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, 0), BRAND_BLUE),
        ("TEXTCOLOR",    (0, 0), (-1, 0), WHITE),
        ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, 0), 8),
        ("ALIGN",        (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME",     (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE",     (0, 1), (-1, 1), 13),
        ("TEXTCOLOR",    (0, 1), (-1, 1), BRAND_BLUE),
        ("BACKGROUND",   (0, 1), (-1, 1), BRAND_LIGHT),
        ("ROWBACKGROUNDS", (0, 1), (-1, 1), [BRAND_LIGHT]),
        ("GRID",         (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("TOPPADDING",   (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 8),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 0.35 * inch))

    # ── Paid-bookings highlight ───────────────────────────────────────────────
    story.append(Paragraph("Booking Details", s["section"]))

    table_header = [
        "Booking ID", "Customer", "Vehicle", "Start", "End", "Days",
        "Amount (KES)", "Status", "Payment"
    ]
    col_widths = [0.9*inch, 1.1*inch, 1.1*inch, 0.7*inch,
                  0.7*inch, 0.4*inch, 0.85*inch, 0.65*inch, 0.65*inch]

    rows = [table_header]
    for b in bookings:
        user    = users_map.get(str(b.get("user_id", "")), {})
        vehicle = vehicles_map.get(str(b.get("vehicle_id", "")), {})
        rows.append([
            str(b["_id"])[-8:],
            (user.get("name") or user.get("email") or "N/A")[:16],
            (vehicle.get("name") or "N/A")[:16],
            b["start_date"].strftime("%d/%m/%y") if b.get("start_date") else "—",
            b["end_date"].strftime("%d/%m/%y")   if b.get("end_date")   else "—",
            str(b.get("days", "—")),
            f"{b.get('total_amount', 0):,.0f}",
            (b.get("status") or "—").title(),
            (b.get("payment_status") or "—").title(),
        ])

    detail_table = Table(rows, colWidths=col_widths, repeatRows=1)

    row_bg = []
    for i in range(2, len(rows)):
        bg = GREY_HEADER if i % 2 == 0 else WHITE
        row_bg.append(("BACKGROUND", (0, i), (-1, i), bg))

    # Status colour coding
    status_colours = {
        "Completed": BRAND_GREEN,
        "Approved":  colors.HexColor("#2563eb"),
        "Pending":   colors.HexColor("#d97706"),
        "Rejected":  colors.HexColor("#dc2626"),
    }
    status_col_idx = 7  # "Status" column index

    detail_table.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, 0), BRAND_BLUE),
        ("TEXTCOLOR",    (0, 0), (-1, 0), WHITE),
        ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, -1), 7.5),
        ("ALIGN",        (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",         (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
        ("TOPPADDING",   (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, GREY_HEADER]),
        ("LINEBELOW",    (0, 0), (-1, 0), 1.5, BRAND_BLUE),
    ] + row_bg))

    story.append(detail_table)
    story.append(Spacer(1, 0.3 * inch))

    # ── Footer note ───────────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=1,
                            color=colors.HexColor("#e2e8f0"), spaceBefore=8))
    story.append(Paragraph(
        "SmartDrive — Kenya's Intelligent Vehicle Hire Platform &nbsp;|&nbsp; "
        "info@smartdrive.co.ke &nbsp;|&nbsp; Westlands, Nairobi &nbsp;|&nbsp; "
        "Confidential — For Internal Use Only",
        s["footer"],
    ))

    doc.build(story)
    buffer.seek(0)
    return buffer


# ── Excel export (multi-sheet) ────────────────────────────────────────────────

def generate_excel_report(bookings: list, vehicles_map: dict, users_map: dict):
    """Generate a multi-sheet Excel workbook. Returns BytesIO."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    wb = Workbook()

    HEADER_FILL = PatternFill("solid", fgColor="1a4d8c")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
    ALT_FILL    = PatternFill("solid", fgColor="F1F5F9")
    BOLD        = Font(bold=True)
    CENTER      = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="CBD5E1")
    BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.fill   = HEADER_FILL
            cell.font   = HEADER_FONT
            cell.alignment = CENTER
            cell.border = BORDER

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)

    # Sheet 1 — Bookings
    ws1 = wb.active
    ws1.title = "Bookings"
    ws1.append(["Booking ID", "Customer", "Email", "Vehicle", "Plate No.",
                 "Start Date", "End Date", "Days", "Amount (KES)",
                 "Status", "Payment Status", "Created At"])
    style_header(ws1)
    for i, b in enumerate(bookings, start=2):
        user    = users_map.get(str(b.get("user_id", "")), {})
        vehicle = vehicles_map.get(str(b.get("vehicle_id", "")), {})
        ws1.append([
            str(b["_id"]),
            user.get("name") or "N/A",
            user.get("email") or "N/A",
            vehicle.get("name") or "N/A",
            vehicle.get("plate_number") or "N/A",
            b["start_date"].strftime("%Y-%m-%d") if b.get("start_date") else "",
            b["end_date"].strftime("%Y-%m-%d")   if b.get("end_date")   else "",
            b.get("days", ""),
            b.get("total_amount", 0),
            b.get("status", ""),
            b.get("payment_status", ""),
            b["created_at"].strftime("%Y-%m-%d %H:%M") if b.get("created_at") else "",
        ])
        if i % 2 == 0:
            for cell in ws1[i]:
                cell.fill = ALT_FILL
    auto_width(ws1)

    # Sheet 2 — Revenue Summary by Status
    ws2 = wb.create_sheet("Revenue Summary")
    ws2.append(["Status", "Count", "Total Amount (KES)"])
    style_header(ws2)
    from collections import defaultdict
    by_status: dict = defaultdict(lambda: {"count": 0, "total": 0.0})
    for b in bookings:
        st = b.get("status", "unknown")
        by_status[st]["count"] += 1
        by_status[st]["total"] += b.get("total_amount", 0)
    for st, data in sorted(by_status.items()):
        ws2.append([st.title(), data["count"], round(data["total"], 2)])
    auto_width(ws2)

    # Sheet 3 — Fleet Performance
    ws3 = wb.create_sheet("Fleet Performance")
    ws3.append(["Vehicle", "Plate No.", "Total Bookings", "Revenue (KES)"])
    style_header(ws3)
    by_vehicle: dict = defaultdict(lambda: {"count": 0, "total": 0.0})
    for b in bookings:
        vid = str(b.get("vehicle_id", ""))
        by_vehicle[vid]["count"] += 1
        by_vehicle[vid]["total"] += b.get("total_amount", 0)
    for vid, data in sorted(by_vehicle.items(), key=lambda x: -x[1]["count"]):
        v = vehicles_map.get(vid, {})
        ws3.append([
            v.get("name") or "Unknown",
            v.get("plate_number") or "N/A",
            data["count"],
            round(data["total"], 2),
        ])
    auto_width(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
